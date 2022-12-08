# -*- coding: utf-8 -*-
"""
Created on Tue Aug  9 12:23:49 2022

@author: user
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
import openpyxl

# 나올 때 까지 기다리는 함수
def wait_until(xpath_str):
    WebDriverWait(browser,30).until(EC.presence_of_element_located((By.XPATH,xpath_str)))
    
browser = webdriver.Chrome(r"C:\Users\user\Desktop\016_\week5\1\chromedriver.exe")
url = "https://flight.naver.com/"
browser.get(url)

begin_date=browser.find_element(By.XPATH, '//button[text()="가는 날"]')
begin_date.click()
now=date.today().month  # 8
count=9-now  # 1
#출발날짜
wait_until('//b[text()="8"]')
day8=browser.find_elements(By.XPATH, '//b[text()="8"]')
day8[count].click()
#도착날짜
wait_until('//b[text()="20"]')
day8=browser.find_elements(By.XPATH, '//b[text()="20"]')
day8[count].click()
#도착지역
wait_until('//b[text()="도착"]')
arrival=browser.find_element(By.XPATH, '//b[text()="도착"]')
arrival.click()
wait_until('//button[text()="유럽"]')
europe=browser.find_element(By.XPATH, '//button[text()="유럽"]')
europe.click()
wait_until('//i[text()="샤를드골국제공항"]')
CDG=browser.find_element(By.XPATH, '//i[text()="샤를드골국제공항"]')
CDG.click()
#직항 선택
wait_until('//button[text()="직항만"]')
direct=browser.find_element(By.XPATH, '//button[text()="직항만"]')
direct.click()
#검색
wait_until('//span[text()="항공권 검색"]')
search=browser.find_element(By.XPATH, '//span[text()="항공권 검색"]')
search.click()

# 자료 탐색
wait_until("//div[@class = 'concurrent_ConcurrentItemContainer__2lQVG result']")
elements = browser.find_elements(By.XPATH, "//div[@class = 'concurrent_ConcurrentItemContainer__2lQVG result']")

# 엑셀 설정
wb = openpyxl.Workbook()
sheet = wb.active
sheet.append(['항공사', '출발시간', '도착시간', '직항여부', '비행시간', '항공사', '출발시간', '도착시간', '직항여부', '비행시간'])

# 데이터 확인 및 저장
for i in elements:
    print(i.text)
    print('==========')
    i1 = i.text.splitlines()
    print(i1)
    sheet.append(i1)
    print('===========')

wb.save(r"C:\Users\user\Desktop\016_\week5\2\fl2.csv")


#wb2 = openpyxl.Workbook()
#sheet = wb2.active
#sheet.cell(row=1, column=1).value='ddd'
#wb2.save(r"C:\Users\user\Desktop\016_\week5\2\fl.csv")

